'''Excel Spreadsheet Utility, mostly Mainlab Chado Loader (MCL) compatible
speadsheet creation.'''

# Spreadsheet writer (xls + xlsx).
import openpyxl
import os
import re

class MCLSpreadsheet():
    '''Writes MCL compatible spreadsheets for chado import.'''

    OUTFILE = 'test.xlsx'

    DB_HEADERS = [
        '*db_name', 'url_prefix', 'url', 'description'
    ]
    CV_HEADERS = [
        '*cv_name', 'definition'
    ]
    CVTERM_HEADERS = [
        '*db_name', '*cv_name', '*cvterm_name', '*accession', 'definition'
    ]
    STOCK_HEADERS = [
        '*stock_name', '*germplasm_type', '*genus', '*species', 'secondary_ID',
        'description', 'subspecies', 'GRIN_ID', 'paternal_parent',
        'maternal_parent', 'mutation_parent', 'selfing_parent', 'alias',
        'cultivar', 'pedigree', 'origin', 'population_size',
        'germplasm_center', 'comments', 'image', 'reference'
    ]
    GEOLOCATION_HEADERS = [
        '*site_name', 'latitude', 'longitude', 'altitude', 'geodetic_datum',
        'type', 'country', 'state', 'region', 'address', 'comments'
    ]
    DATASET_HEADERS = [
        '*dataset_name', '*type', 'sub_type', 'super_dataset',
        'trait_descriptor_set', 'PI', 'crop', 'comments', 'reference',
        'permission', 'description'
    ]
    CONTACT_HEADERS = [
        '*contact_name', 'alias', '*type', 'first_name', 'last_name',
        'institution', 'lab', 'address', 'email', 'phone', 'fax', 'title',
        'country', 'source', 'last_update', 'url', 'comments', 'keywords'
    ]
    PHENO_HEADERS = [
        '*dataset_name', '*stock_name', '*genus', '*species', '*sample_ID',
        'clone_ID', 'evaluator', 'site_name', 'rep', 'rootstock', 'plot',
        'row', 'position', 'plant_date', 'data_year', 'evaluation_date',
        'pick_date', 'fiber_pkg', 'storage_time', 'storage_regime', 'comments'
    ]

    def __init__(self, chado_connection):
        '''Expecting an initialized database cursor to the chado schema.'''
        self.created_xlsxs = []
        self.chado = chado_connection

    def __create_xlsx(self, headers, content, title=''):
        '''Create a spreadsheet like this:
        
        N, M = len(headers), len(content)

             headers[0]     | headers[..]     | headers[N]
            ----------------+-----------------+---------------
             content[0][0]  | content[0][..]  | content[0][N]
             content[..][0] | content[..][..] | content[..][N]
             content[M][0]  | content[M][..]  | content[M][N]

        Note: The type(content[n]) may be a list() or a dict(), in later case
              values are assigned to the columns indicated by the keys
              (numbers/letters).
        '''
        wb = openpyxl.Workbook()
        ws1 = wb.get_active_sheet()
        if title: ws1.title = title

        for n, header in enumerate(headers):
            ws1.cell(row=0, column=n).value = header

        for row in content:
            if type(row) == dict:
                row = self.__how_i_understand_dict(ws1, row)
            ws1.append(row)

        self.created_xlsxs.append(wb)
        return wb

    def __how_i_understand_dict(self, s, d):
        '''A change to the openpyxl.worksheet.append function.

        Makes dictionary keys refer to row-0 content, not only coordinate. If a
        key-matching row-0 column does not exist, we create it.
        '''
        headers = [h.value for h in s.rows[0]]
        new_d = {}

        for key,value in d.iteritems():
            if key in headers:
                key = headers.index(key)
            else:
                s.cell(row=0, column=len(headers)).value = key
                headers = [h.value for h in s.rows[0]] # needed for len()
                key = headers.index(key)
            new_d[key] = value

        return new_d

    def create_TYPE(self, filename, content, TYPE, sheetname=''):
        '''Create, save, and return a <TYPE> spreadsheet with content.

        If sheetname is set, the sheetname is changed to sheetname.
        The TYPE is an array used as header. We supply some MCL defaults:
            self.DB_HEADERS
            self.CV_HEADERS
            self. ..

        Note: The type(content[0]) may be a list() or a dict(), in later case
              values are assigned to the columns indicated by the keys
              (numbers/keywords).
        '''
        wb = self.__create_xlsx(TYPE, content)
        if sheetname:
            wb.get_active_sheet().title = sheetname
        if not os.path.exists(os.path.dirname(filename)):
            os.makedirs(os.path.dirname(filename))
        wb.save(filename)
        del wb # XXX need to free memory.. 
        return filename

    def create_db(self, filename, name, description=''):
        '''Convenience wrapper around create_TYPE()'''
        content = [[name, '', '', description]]
        return self.create_TYPE(filename, content, self.DB_HEADERS, 'db')

    def create_cv(self, filename, name, description=''):
        '''Convenience wrapper around create_TYPE()'''
        content = [[name, description]]
        return self.create_TYPE(filename, content, self.CV_HEADERS, 'cv')

    def create_cvterm(self, filename, dbname, cvname, cvterm, accession=[],
        definition=[]):
        '''Convenience wrapper around create_TYPE()
        
        By default accessesion will be set equal to cvterm and definition will
        be empty.

        Arguments:                  Types:
            dbname, cvname              str
            cvterm,accession,defi..     [cvterm_name1, cvterm_name2, ..]
        '''
        if not accession:
            accession = cvterm

        if len(cvterm) != len(accession):
            raise RuntimeError('[.create_cvterm] argument length unequal!')
        if definition:
            if len(cvterm) != len(accession) or len(cvterm) != len(definition):
                raise RuntimeError('[.create_cvterm] argument length unequal!')

        if not definition:
            it = zip(cvterm, accession)
            content = [[dbname, cvname, c, a, ''] for c,a in it]
        else:
            it = zip(cvterm, accession, definition)
            content = [[dbname, cvname, c, a, d] for c,a,d in it]

        return self.create_TYPE(filename, content, self.CVTERM_HEADERS,
                                'cvterm')

    # We need this function, but MCL has no template for it => chado.py
    #def create_*ORGANISM*():
    def create_organism(self, genus, species, abbreviation='', common_name='', comment=''):
        '''Create an organism.
        
        Note that this function directly accesses the Chado underlying postgresDB.
        '''
        if chado.has_species(species):
            msg = 'Tried to create an existing species: {}'.format(species)
            raise RuntimeError(msg)
        if chado.has_genus(genus):
            msg = 'Tried to create an existing genus: {}'.format(genus)
            raise RuntimeError(msg)
        chado.create_organism(genus, species, abbreviation, common_name, comment)

    def create_stock(self, filename, name, germplasm_type, genus, species):
        '''Convenience wrapper around create_TYPE()
        
        The 'secondary_ID' will be set equal to name (germplasm name).

        Given a <name> argument as a list of list()'s not a list of strings, we
        interpret the first[0] entry as the uniquename and the second as the
        name.
        '''
        if len(name[0]) == 2:
            # given a list() of list()'s, the first element
            # XXX we discard the uniquename because its just some weird number
            # in our Oracle DB
            content = [[g, germplasm_type, genus, species, g] for n,g in name]
        else:
            content = [[g, germplasm_type, genus, species, g] for g in name]
        return self.create_TYPE(filename, content, self.STOCK_HEADERS, 'stock')

    def create_dataset(self, filename, dataset_name, type, sub='', super=''):
        '''Convenience wrapper around create_TYPE()
        
        A Project/Dataset/Experiment, or just a set of data.
        '''
        content = [[dataset_name, type, sub, super]]
        return self.create_TYPE(filename, content, self.STOCK_HEADERS,
                                'dataset')

    def create_geolocation(self, filename, names, alts, lats, longs,
        fmt='NSEW'):
        '''Convenience wrapper around create_TYPE()
        
        Create a nd_geolocation entry.
        The coordinates must be either given with a '+' or a '-' sign,
        indicating NS, and EW; or appended NSEW, depinding on the fmt=
        argument, which is either '+-' or 'NSEW'.
        '''
        if not fmt in ['NSEW', '+-']:
            raise RuntimeError('[.create_geolocation] fmt wrong')

        # Strip leading whitespace, leading zeroes, and substitute N E S W..
        if fmt == 'NSEW':
            for crdl in [alts, lats, longs]:
                for crd in crdl:
                    try:
                        idx = crdl.index(crd)
                        crdl[idx] = re.sub(r'^\s*0*', '', crdl[idx])
                        crdl[idx] = re.sub(r'([0-9]*)[NE]', '+\\1', crdl[idx])
                        crdl[idx] = re.sub(r'([0-9]*)[SW]', '-\\1', crdl[idx])
                    except TypeError, ValueError:
                        # Either we found a plain int() or the first
                        # substitution was already successfull, and the second
                        # fails, which is both fine.
                        continue

        it = zip(names, alts, lats, longs)
        content = [[na, al, la, lo] for na,al,la,lo in it]

        return self.create_TYPE(filename, content, self.GEOLOCATION_HEADERS,
                                'site')

    def create_contact(self, filename, names, types, optionals=None):
        '''Convenience wrapper around create_TYPE()
        
        Create one/manny contact entry(s). <optionals> must be a list() of
        dict()s with keys element self.CONTACT_HEADERS, except 'contact_name'
        and 'type' as those are already required in names and types.
        '''
        content = []
        for n,t,opt_d  in zip(names, types, optionals):
            d = {}
            d['*contact_name'] = n
            d['*type'] = t
            d.update(opt_d)
            content.append(d)
        return self.create_TYPE(filename, content, self.CONTACT_HEADERS, 'contact')

    def create_phenotype(self, filename, dataset_name, stocks, descriptors,
        other=[], genus='', species='', sample_id='', clone_id='', contact=''):
        '''Upload Phenotype Data.
        
        Arguments:
            stocks      A list of stock values, which MUST already be in the
                        database.
            genus       Might be omitted if species is specified.
            species     Might be omitted if genus is specified.
            descriptors An array of dict()s, where the keys are column names
                        and MUST start with a '#', and the values .. are values
            other       another array of dict()s like ^
                        with keys e{ self.PHENO_HEADERS[5:] }
            sample_id   If omitted, will be contructed from key and value of
                        the descriptors dict items
            clone_id    CAN be omitted, or specified via 'other'
            contact     CAN be omitted, or specified via 'other'
        '''
        # Disambiguate the organism if necessary.
        if not genus:
            where_arg = "species = '{}'".format(species)
            orga = self.chado.get_organism(where=where_arg)
            if len(orga) == 1:
                genus = orga[0].genus
        if not species:
            where_arg = "genus = '{}'".format(genus)
            orga = self.chado.get_organism(where=where_arg)
            if len(orga) == 1:
                species = orga[0].species
        if not genus and not species:
            msg = 'Could not disambiguate organism, need genus, species or both'
            raise RuntimeError(msg)

        for ds in descriptors:
            for d in ds.keys():
                if '#' != d[0]:
                    msg = 'phenotype descriptor({0}->{1}) must begin with "#"'
                    class StupidUserError(RuntimeError):
                        pass
                    raise StupidUserError(msg.format(d, ds[d]))
                if len(d) < 3:
                    msg = 'phenotype descriptor({0}->{1}) cannot be empty'
                    raise RuntimeError(msg.format(d, ds[d]))

        content = []
        if other:
            for descs, oths, stock in zip(descriptors, other, stocks):
                sid = '{0}_{1}'.format(descs.keys()[0][1:], descs.values()[0])
                c_dict = {'*dataset_name':dataset_name, '*stock_name':stock,
                    '*genus':genus, '*species':species, '*sample_ID':sid,
                    'clone_ID':clone_id, 'evaluator':contact}
                c_dict.update(descs)
                c_dict.update(oths)
                content.append(c_dict)
        else:
            for descs, stock in zip(descriptors, stocks):
                sid = '{0}_{1}'.format(descs.keys()[0][1:], descs.values()[0])
                c_dict = {'*dataset_name':dataset_name, '*stock_name':stock,
                    '*genus':genus, '*species':species, '*sample_ID':sid,
                    'clone_ID':clone_id, 'evaluator':contact}
                c_dict.update(descs)
                content.append(c_dict)

        return self.create_TYPE(filename, content, self.PHENO_HEADERS,
                                'phenotype')

